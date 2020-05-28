'''
numbers=[5,2,5,2,2]

for item in numbers:
    print('x' * item)


for item in numbers:
    # reset to empty each time
    output=""
    # range of 5 (1,2,3,4)
    for item in range(item):
        output += 'x'
    print(output)


-------------------------------------------------------
'''

# list_1=[1,2,3,4,5,6,7,8]

# print(max(list_1))

# maximum=list_1[0]
# for i in list_1:
#     if i > maximum:
#         maximum=i
# print(maximum)
    

'''
class Person:
    # argument/attribute
    def __init__(self,name):
        self.name=name
    # method
    def talk(self):
        print("talk")

john=Person("John Smith")
print(john.name)
# Output: 
#John Smith
john.talk()
# calling the method
# Prints the talk
'''

'''
# Adding a f" so that we can use a variable
class Person:
    # argument/attribute
    def __init__(self,name):
        self.name=name
    # method
    def talk(self):
        print(f"Hi I am {self.name}")
        # Now the name function can be called using the talk method.


john=Person("John Smith")
john.talk()
# 
# calling the method
# Prints the talk function
bob = Person("Bob Smith"):
# each object is a different instance in the person class.
# creates a new instances in the Person class.
bob.talk()
# We have to call bob in order to print him
'''


'''
#Inheritance = a mechanism of reusing code.
class Mammal:
    def walk(self):
        print("walk")
class Dog(Mammal):
    def bark(self):
        print("bark")
    # dog will inherit everything from the parent class (Mammal)
    #inherit all methods defined in Mammal
    # Python does not like empty classes. Use pass to tell python to ignore it
       

# if we have to repeat the code then the code is bad.
#DRY = Don't Repeat Yourself.

class Cat(Mammal):
   def be_annoying(self):
       print("annoying")



#creating objects:
dog1=Dog()
dog1.walk()

cat1=Cat()
cat1.be_annoying
'''
'''
#Modules: file with python code that is used to organize code
#Cut the below code and put it into a new file called converters.py (name can vary)
#Keep any code that is related to converting on that file
def lbs_to_kg(weight):
    return weight * 0.45


def kg_to_lbs(weight):
    return weight/0.45



# importing converters(module)
import converters 
#now it is an object in our old file
from converters import kg_to_lbs
#allows you to call a specific function in the module
kg_to_lbs(100)
#we can call function without prefixing it

converters.kg_to_lbs(70)
# call the functions and pass 70kg
'''
'''
#Organizing codes properly into functions modules and classes.




#take the function below and place into a new document.

def find_max(list_1):
    maximum=list_1[0]

    for i in list_1:
        if i > maximum:
            maximum=i
    return maximum



from utils import find_max
list_1=[1,2,3,4,5,6,7,8]
maximum1=find_max(list_1)
# When you change the meaning of the function it is bad

print(maximum1(list_1))

# shift f6 for renaming and changing
'''
'''
#Packages=another way to organize code
#organize related modules into a package
#A directory or a folder


1. right click and create new directory.
    call it ecommerce
2.right click directory and add new file
    __init__ #and save file
    #it will treat the directory as a package
1.Right click project
2. Create New Python Package
3. Name it
4.Create module (def calc_shipping())


---------------
# Great for Django, each package typically contains several modules
import.packagename.filename
packagename.filename(module).function

package=ecommerce
filename=shipping
calc_shipping=function

from ecommerce.shipping import calc_shipping, cal_tax, etc


from ecommerce import shipping
#now we can access all modules using
shipping.calc_shipping()
shipping.calc_tax()
------------------------------
#Research common modules(func) to know for AI and as a software developer
import random
#stored in external libraries python 3.8 library root
#contains al of the buit in modules in python
#shows you the source code for the modue

Random Values

random.random()
for i in range(3):
    print(random.random())
#generates a random value between 0 to 1
#
#
#
random.randint()
for i in range(3):
    print(random.randint(10,20))
#random value for a particular range
#pass two arguments to specify the range
#
#
random.choice()
members=["John", "Mary", "Haoua"]
print(random.choice(members))
#picking a random choice in a list
#
#
#
#
'''
'''
Dice
- roll() tuple of 2 random values
'''
'''
import random


class Dice:
    
    def roll(self):
        
        first=random.randint(1,6)
        second=random.randint(1,6)
        return first, second
        #return x, y also returns a tuple

dice = Dice()
print(dice.roll())
'''
#Directories
pathlib = module that

from pathlib import Path
#now create path object

#absolute path: path starting from the root of our hardisk
    #c:\Program Files\Microsoft
#relative path: local path


path=Path("ecommerce")
#if we do not put anything in it it will default to the current directory.
print(path.exists())
#Output=False because I have no create an eccommerce package

path=Path("emails")
print(path.mkdir())
#.mkdir() creates a new Package
print(path.rmdir())
#removes the directory
------------------
path= Path()
#first change the path to the current directory
print(path.glob('*'))
#Astrict means everything, all files and directories
#We can find all of the current directories and paths
print(path.glob('*.*'))
#gives us all of the files but not the directories
print(path.glob('*.py'))
#gives us all of the Python directories
#gives us a generator object (we can itirate or loop through them)
for file in path.glob('*.py')
PRINT(FILE)
#will print the files create in the current path
for file in (path.glob('*')):
    print(file)
    #gives us a list of files and directories
----------------------------------------------
#Pyip pip
#Python package index or pypi.org
#published packages for certain functionality.
#not all are debugged
#Webscrapping = have engine browse website and extract information
#Engines are webcrawlers extract information and find the title of your post
#browser automation = selenium or celenium

#How to install packages from Pypi.org
1. Search for package
2.Type  pip install openpyxl
#pip is used to install and unstall packages on pypi.org
3.hit enter
-----------------------
#Write a Python Program that can process thousands of spreadsheets:
1.Right Click Project
2. Click Reveal in Finder or Reveal in Explorer:
#Opens up directory where project is stored
3. import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')
wb['Sheet1'] = sheet
#returns our first sheet

#These two return the same values
sheet['a1']
#accessing cells
#type coordinates of the cell (col.A row1)
sheet.cell(1,1) = cell



print(cell.value)
#transaction_id is the output (it is the value in cell a1)




1.print(sheet.max_row)
#max_row is an attribute in sheet
#we get 4 because we have a total of 4 rows


for row in range(1,sheet.max_row + 1)
#add one to include four
PRINT(ROW)


for row in range(2,sheet.max_row + 1)
#ignore first row because it is heading
cell=sheet.cell(row,3)

sheet.cell #gives us access to the cell at this row and the column is 3

PRINT(cell.value)


=============
#add corrected prices in new column:
for row in range(2,sheet.max_row + 1)
    cell=sheet.cell(row,3)
    corrected_price= cell.value * 0.9
    sheet.cell(row, 4) = corrected_price_cell
    #cell object in the spreadsheet
    corrected_price_cell.value=corrected_price
wb.save("transactions2.xlsx")
#saves new file



====
#adding a chart
import openpyxl as xl

from openpyxl.chart import BarChart, Reference
#module chart, and we are importing two classes (first letter of everyword capalitized)


for row in range(2,sheet.max_row + 1)
    cell=sheet.cell(row,3)
    corrected_price= cell.value * 0.9
    sheet.cell(row, 4) = corrected_price_cell
    #cell object in the spreadsheet
    corrected_price_cell.value=corrected_price
wb.save("transactions2.xlsx")
#after the for loop
values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4
max_col=4)
#select rows from two to four
chart=BarChart()
chart.add_data(values)
#add our values
sheet.add_chart(chart,"e2")
#e2 is here the top left corner of the chart will be
#Open py excel


======
def process_workbook(filename):
    wb=xl.load_workbook(filename)
    wb['Sheet1'] = sheet
#How to use on any values
    for row in range(2,sheet.max_row + 1)
        cell=sheet.cell(row,3)
        corrected_price= cell.value * 0.9
        sheet.cell(row, 4) = corrected_price_cell
        #cell object in the spreadsheet
        corrected_price_cell.value=corrected_price
        #after the for loop
        values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4
        max_col=4)
        #select rows from two to four
        chart=BarChart()
        chart.add_data(values)
        #add our values
        sheet.add_chart(chart,"e2")
        #e2 is here the top left corner of the chart will be
        #Open py excel
        wb.save(filename)

=====
#Machine Learning: Steps
1. Import data (often in form of csv file)
2. Clean the data (remove duplicates)
3. Split the Data into Training/Test 
4. Create a model : select algorithm for data (decision trees etc)
#depends on problem Sci kid learn is one of the more popular ones
5.Training model
6.Make Predictions
7.Evaluate and Improve

#Machine Learning: Libraries
1.Numpy=very popular
2.Pandas=Data frame (row, columns, popular for machie learning and data science)
3.MatPlotLib=
4.Scikit-Learn=

Jupyter = best coding platform for machine learning; makes it easy to test data







